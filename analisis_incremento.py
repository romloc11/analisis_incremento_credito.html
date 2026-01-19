import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment
from openpyxl.chart import PieChart, BarChart, Reference


def motor_credito_ciosa_vFinal(ruta_excel):

    # ==========================================================
    # 1. CARGA DE DATOS
    # ==========================================================
    if not os.path.exists(ruta_excel):
        print(f"Error: No se encontró el archivo en: {ruta_excel}")
        return

    print("Cargando y limpiando datos...")
    df = pd.read_excel(ruta_excel)
    df.columns = df.columns.astype(str).str.strip()

    # ----------------------------------------------------------
    # NORMALIZACIÓN CLIENTE (NO BORRAR TODAVÍA)
    # ----------------------------------------------------------
    df['Cliente_norm'] = (
        df['Cliente']
        .astype(str)
        .str.replace('.0', '', regex=False)
        .str.strip()
        .str.lstrip('0')
    )

    # ==========================================================
    # 2. HISTORIAL DE MODIFICACIONES DE LÍMITE
    # ==========================================================
    print("Cargando historial de modificaciones de límite...")

    df_limite = pd.read_excel(
        ruta_excel,
        sheet_name='incremento'
    )

    df_limite.columns = df_limite.columns.astype(str).str.strip()

    df_limite['CodigoSAP_norm'] = (
        df_limite['Código SAP']
        .fillna('')
        .astype(str)
        .str.replace('.0', '', regex=False)
        .str.strip()
        .str.lstrip('0')
    )

    df_limite['FechaModificacion'] = pd.to_datetime(
        df_limite['Historial de aprobaciones/Fecha de resolución'],
        errors='coerce'
    )

    df_limite = (
        df_limite
        .sort_values('FechaModificacion', ascending=False)
        .drop_duplicates('CodigoSAP_norm')
    )

    # ----------------------------------------------------------
    # CRUCE HISTORIAL
    # ----------------------------------------------------------
    df = df.merge(
        df_limite[['CodigoSAP_norm', 'FechaModificacion']],
        left_on='Cliente_norm',
        right_on='CodigoSAP_norm',
        how='left'
    )

    df.drop(columns=['CodigoSAP_norm'], inplace=True)

    df['Ultima_Modificacion_Limite'] = df['FechaModificacion'].dt.date

    hoy = pd.Timestamp.today()

    df['Dias_Desde_Modificacion'] = (
        hoy - df['FechaModificacion']
    ).dt.days

    df['Recientemente_Modificado'] = (
        df['Dias_Desde_Modificacion'].notna()
        & (df['Dias_Desde_Modificacion'] <= 90)
    )

    # ==========================================================
    # 3. INFORMACIÓN DE COBERTURA
    # ==========================================================
    print("Cargando información de cobertura...")

    df_cobertura = pd.read_excel(
        ruta_excel,
        sheet_name='cobertura'
    )

    df_cobertura.columns = df_cobertura.columns.astype(str).str.strip()

    df_cobertura['CodigoSAP_norm'] = (
        df_cobertura['Name']
        .fillna('')
        .astype(str)
        .str.replace('.0', '', regex=False)
        .str.strip()
        .str.lstrip('0')
    )

    # ----------------------------------------------------------
    # CRUCE COBERTURA
    # ----------------------------------------------------------
    df = df.merge(
        df_cobertura[
            ['CodigoSAP_norm', 'PAGARE', 'CONTRATO', 'INE TITULAR/REPRESENTANTE']
        ],
        left_on='Cliente_norm',
        right_on='CodigoSAP_norm',
        how='left'
    )

    df.drop(columns=['CodigoSAP_norm'], inplace=True)

    # ==========================================================
    # 4. FUNCIONES DE PUNTUACIÓN
    # ==========================================================
    def calcular_pts_uso(row):
        limite = pd.to_numeric(row.get('Límite de crédito', 1), errors='coerce')
        limite = 1 if pd.isna(limite) or limite <= 0 else limite

        meses = int(
            max(1, min(pd.to_numeric(row.get('MesesAntiguedad', 1), errors='coerce'), 12))
        )

        ventas = [
            pd.to_numeric(row.get(f'VtaMes{i}', 0), errors='coerce') or 0
            for i in range(1, meses + 1)
        ]

        promedio_venta = np.mean(ventas)
        uso = promedio_venta / limite

        return min(max(uso * 10, 0), 10)

    def calcular_pts_adn(row):
        raw_adn = row.get('ClasificacionActual', 'DDD')

        adn = (
            str(raw_adn).upper()
            .replace(" ", "")
            .replace("-", "")
            .replace("/", "")
            .replace("\n", "")
            .strip()
        )

        if len(adn) != 3:
            adn = "DDD"

        mapeo = {'A': 10, 'B': 7, 'C': 4, 'D': 0}

        pts_pago = mapeo.get(adn[2], 0)
        pts_compra = mapeo.get(adn[0], 0)

        return (pts_pago * 0.7) + (pts_compra * 0.3)

    def calcular_pts_variabilidad(row):
        meses = int(
            max(3, min(pd.to_numeric(row.get('MesesAntiguedad', 1), errors='coerce'), 12))
        )

        ventas = [
            pd.to_numeric(row.get(f'VtaMes{i}', 0), errors='coerce') or 0
            for i in range(1, meses + 1)
        ]

        prom = np.mean(ventas)
        if prom <= 0:
            return 0

        cv = np.std(ventas) / prom
        return max(0, 10 - (cv * 10))

    def calcular_pts_capacidad_pago(row):
        limite = pd.to_numeric(row.get('Límite de crédito', 1), errors='coerce')
        limite = 1 if pd.isna(limite) or limite <= 0 else limite

        pago_max = pd.to_numeric(row.get('PagosMaximo', 0), errors='coerce') or 0
        ratio = pago_max / limite

        if ratio >= 1:
            return 10
        elif ratio >= 0.75:
            return 8
        elif ratio >= 0.50:
            return 6
        elif ratio >= 0.30:
            return 3
        else:
            return 0

    # ==========================================================
    # 5. CÁLCULO DE SCORES
    # ==========================================================
    print("Calculando Score de Riesgo...")

    df['Pts_Uso'] = df.apply(calcular_pts_uso, axis=1)
    df['Pts_ADN'] = df.apply(calcular_pts_adn, axis=1)
    df['Pts_Variabilidad'] = df.apply(calcular_pts_variabilidad, axis=1)

    df['Pts_DPP'] = pd.to_numeric(df['DPPPonderado'], errors='coerce').apply(
        lambda x: 10 if x <= 0 else (7 if x <= 7 else (4 if x <= 15 else 0))
    )

    df['Pts_Antiguedad'] = pd.to_numeric(df['MesesAntiguedad'], errors='coerce').apply(
        lambda x: 10 if x >= 24 else (7 if x >= 12 else 4)
    )

    df['Pts_Vencido'] = df.apply(
        lambda r: 10
        if pd.to_numeric(r.get('%SaldoVencido', 0), errors='coerce') == 0
        and pd.to_numeric(r.get('DiasMasVencida', 0), errors='coerce') == 0
        else 0,
        axis=1
    )

    df['Pts_CapacidadPago'] = df.apply(calcular_pts_capacidad_pago, axis=1)

    pesos = {
        'Pts_Uso': 0.47,
        'Pts_ADN': 0.10,
        'Pts_DPP': 0.13,
        'Pts_Vencido': 0.12,
        'Pts_Antiguedad': 0.05,
        'Pts_Variabilidad': 0.06,
        'Pts_CapacidadPago': 0.10
    }

    df['SCORE_FINAL'] = round(sum(df[col] * peso for col, peso in pesos.items()) * 10)

    # ==========================================================
    # 5.1 DECISIÓN DE CRÉDITO SEGÚN SCORE
    # ==========================================================
    def decision_credito(row):

        score = row.get('SCORE_FINAL')
        clasificacion = str(row.get('ClasificacionActual', '')).strip().upper()

        # --------------------------------------------------
        # REGLA DURA: CLASIFICACIÓN ESPECIAL
        # --------------------------------------------------
        if clasificacion in ["N", "X", "XL"]:
            return "Sin cambio"

        # --------------------------------------------------
        # VALIDACIÓN SCORE
        # --------------------------------------------------
        if pd.isna(score):
            return "Sin información"
        elif score >= 90:
            return "Incremento"
        elif 80 <= score < 90:
            return "Posible incremento (Revisión manual)"
        elif 20 <= score < 80:
            return "Sin cambio"
        elif 10 <= score < 20:
            return "Posible decremento (Revisión manual)"
        else:
            return "Decremento"
        
    df['Decision_Credito'] = df.apply(decision_credito, axis=1)

    # ==========================================================
    # 5.2 MONTO SUGERIDO DE CRÉDITO
    # ==========================================================
    def calcular_monto_sugerido(row):

        limite_actual = pd.to_numeric(row.get('Límite de crédito', 0), errors='coerce')
        limite_actual = 0 if pd.isna(limite_actual) else limite_actual

        cartera_total = pd.to_numeric(row.get('Cartera total', 0), errors='coerce')
        cartera_total = 0 if pd.isna(cartera_total) else cartera_total

        score = pd.to_numeric(row.get('SCORE_FINAL', 0), errors='coerce')
        score = 0 if pd.isna(score) else score

        decision = row.get('Decision_Credito', 'Sin cambio')

        # ----------------------------
        # VENTAS (12 MESES)
        # ----------------------------
        ventas = [
            pd.to_numeric(row.get(f'VtaMes{i}', 0), errors='coerce') or 0
            for i in range(1, 13)
        ]

        prom_ventas = np.mean(ventas)
        max_ventas = max(ventas) if ventas else 0

        # Base operativa mínima
        base_operativa = max(prom_ventas, max_ventas * 0.75)

        # ----------------------------
        # CÁLCULO SEGÚN DECISIÓN
        # ----------------------------
        if decision == "Incremento":
            # factor dinámico por score (90 → 1.15, 100 → 1.35)
            factor = 1.15 + ((score - 90) / 10) * 0.20
            sugerido = max(base_operativa / 0.45, limite_actual * factor)

            # límite de crecimiento máximo
            sugerido = min(sugerido, limite_actual * 1.40)

        elif decision == "Posible incremento (Revisión manual)":
            factor = 1.05 + ((score - 80) / 10) * 0.10
            sugerido = max(base_operativa / 0.60, limite_actual * factor)

            sugerido = min(sugerido, limite_actual * 1.25)

        elif decision == "Posible decremento (Revisión manual)":
            sugerido = max(base_operativa * 0.95, limite_actual * 0.85)

        elif decision == "Decremento":
            sugerido = max(base_operativa * 0.80, limite_actual * 0.70)

        else:  # Sin cambio
            sugerido = limite_actual

        # ----------------------------
        # PROTECCIONES FINALES
        # ----------------------------

        # Nunca menor a cartera vigente
        sugerido = max(sugerido, cartera_total)

        # Nunca menor a 1 mes de ventas promedio
        sugerido = max(sugerido, prom_ventas)

        # Redondeo a múltiplos de 5,000
        sugerido = round(sugerido / 5000) * 5000

        return sugerido

    df['Monto_Sugerido_Credito'] = df.apply(calcular_monto_sugerido, axis=1)

    # ==========================================================
    # 6. LIMPIEZA FINAL (AHORA SÍ)
    # ==========================================================
    df.drop(columns=['Cliente_norm'], inplace=True)

     # ==========================================================
    # 7. EXPORTACIÓN CON FORMATO Y CONTROL DE ERRORES
    # ==========================================================
    ruta_salida = os.path.join(
        os.path.dirname(ruta_excel),
        "Resultado_Analisis_credito.xlsx"
    )

    try:
        # ----------------------------
        # ESCRITURA DEL EXCEL
        # ----------------------------
        df.to_excel(ruta_salida, index=False)

    except PermissionError:
        print("\n❌ ERROR DE PERMISOS")
        print("El archivo de salida está abierto o no tienes permisos:")
        print(ruta_salida)
        print("➡️ Cierra el archivo y vuelve a ejecutar el script.")
        return

    except Exception as e:
        print("\n❌ ERROR AL EXPORTAR ARCHIVO")
        print(f"Detalle: {e}")
        return

    try:
        # ----------------------------
        # FORMATO DE TABLA Y COLUMNAS
        # ----------------------------
        wb = load_workbook(ruta_salida)
        ws = wb.active
        ws.title = "Evaluacion_Crediticia"

        # Rango completo
        max_row = ws.max_row
        max_col = ws.max_column
        col_letter_end = ws.cell(row=1, column=max_col).column_letter

        tabla = Table(
            displayName="AnalisisCredito",
            ref=f"A1:{col_letter_end}{max_row}"
        )

        estilo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)

        # ----------------------------
        # AJUSTE AUTOMÁTICO DE COLUMNAS
        # ----------------------------
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = min(max_length + 2, 45)

   


        # ==========================================================
        # HOJA RESUMEN EJECUTIVO
        # ==========================================================
        ws_resumen = wb.create_sheet("Resumen_Ejecutivo")

        ws_resumen["A1"] = "Resumen Ejecutivo de Crédito"
        ws_resumen["A1"].font = Font(bold=True, size=14)

        resumen = {
            "Total clientes": len(df),
            "Score promedio": round(df["SCORE_FINAL"].mean(), 1),
            "Incrementos": (df["Decision_Credito"] == "Incremento").sum(),
            "Posibles incrementos": df["Decision_Credito"].str.contains("Posible incremento").sum(),
            "Sin cambio": (df["Decision_Credito"] == "Sin cambio").sum(),
            "Posibles decrementos": df["Decision_Credito"].str.contains("Posible decremento").sum(),
            "Decrementos": (df["Decision_Credito"] == "Decremento").sum(),
            "Límite actual total": df["Límite de crédito"].sum(),
            "Límite sugerido total": df["Monto_Sugerido_Credito"].sum(),
            "Impacto neto": df["Monto_Sugerido_Credito"].sum() - df["Límite de crédito"].sum()
        }

        fila = 3
        for k, v in resumen.items():
            ws_resumen[f"A{fila}"] = k
            ws_resumen[f"B{fila}"] = v
            fila += 1

        ws_resumen.column_dimensions["A"].width = 35
        ws_resumen.column_dimensions["B"].width = 20
     
    except PermissionError:
        print("\n❌ ERROR DE PERMISOS")
        print("El archivo de salida se abrió mientras se aplicaba el formato:")
        print(ruta_salida)
        print("➡️ Ciérralo y ejecuta nuevamente.")
        return

    except Exception as e:
        print("\n❌ ERROR AL FORMATEAR EL ARCHIVO")
        print(f"Detalle: {e}")
        return

    print("\n✅ Proceso exitoso")
    print(f"📁 Archivo generado: {ruta_salida}")

    # ===============================
    # KPIs VISUALES
    # ===============================
    kpis = [
        ("Total Clientes", len(df)),
        ("Incrementos", (df["Decision_Credito"] == "Incremento").sum()),
        ("Decrementos", (df["Decision_Credito"] == "Decremento").sum()),
        ("Score Promedio", round(df["SCORE_FINAL"].mean(), 1)),
        ("Impacto Neto", resumen["Impacto neto"])
    ]

    col = 4
    for titulo, valor in kpis:
        ws_resumen.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col+1)
        celda = ws_resumen.cell(row=2, column=col)
        celda.value = f"{titulo}\n{valor:,.0f}" if isinstance(valor, (int, float)) else f"{titulo}\n{valor}"
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.fill = PatternFill("solid", fgColor="BDD7EE")
        col += 2

    # ===============================
    # GRÁFICA PIE - DECISIONES
    # ===============================
    fila_inicio = fila + 2

    decisiones = df["Decision_Credito"].value_counts()

    ws_resumen[f"A{fila_inicio}"] = "Decisión"
    ws_resumen[f"B{fila_inicio}"] = "Cantidad"

    for i, (dec, cnt) in enumerate(decisiones.items(), start=fila_inicio + 1):
        ws_resumen[f"A{i}"] = dec
        ws_resumen[f"B{i}"] = cnt

    pie = PieChart()
    pie.title = "Distribución de Decisiones de Crédito"

    labels = Reference(ws_resumen, min_col=1, min_row=fila_inicio + 1, max_row=i)
    data = Reference(ws_resumen, min_col=2, min_row=fila_inicio, max_row=i)

    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    ws_resumen.add_chart(pie, "D8")

    # ===============================
    # GRÁFICA BARRAS - LIMITES
    # ===============================
    fila_lim = i + 3

    ws_resumen[f"A{fila_lim}"] = "Concepto"
    ws_resumen[f"B{fila_lim}"] = "Monto"

    ws_resumen[f"A{fila_lim+1}"] = "Límite Actual"
    ws_resumen[f"B{fila_lim+1}"] = resumen["Límite actual total"]

    ws_resumen[f"A{fila_lim+2}"] = "Límite Sugerido"
    ws_resumen[f"B{fila_lim+2}"] = resumen["Límite sugerido total"]

    bar = BarChart()
    bar.title = "Impacto en Límites de Crédito"
    bar.y_axis.title = "Monto"
    bar.x_axis.title = "Tipo"

    data = Reference(ws_resumen, min_col=2, min_row=fila_lim, max_row=fila_lim+2)
    cats = Reference(ws_resumen, min_col=1, min_row=fila_lim+1, max_row=fila_lim+2)

    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)

    ws_resumen.add_chart(bar, "D24")
    wb.save(ruta_salida)


if __name__ == "__main__":

    ruta = input("Ruta del archivo a analizar: ")
    motor_credito_ciosa_vFinal(ruta)
